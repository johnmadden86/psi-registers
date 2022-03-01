from datetime import datetime

from openpyxl import load_workbook
from scrape import write_to_json

file_path = 'C:/Users/johnm/Downloads/Copy of SISV data as at 4th Feb.xlsx'
wb = load_workbook(filename=file_path)
for sheet in wb:
    keys = list(sheet.values)[0]
    obj_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        obj = {}
        for i, r in enumerate(row):
            try:
                r = datetime.strftime(r, '%Y-%m-%d')
            except TypeError:
                pass
            obj[keys[i]] = r
        obj_list.append(obj)
    json_file = f'{sheet.title}.json'
    write_to_json(obj_list, json_file)

